__author__ = 'Joseph Conlin'
"""
Classes for outputting movie information in various formats
"""
from TheaterDetailPage import TheaterDetail, MovieDetail

import os
import csv
# import xlwt
from openpyxl import Workbook, load_workbook
import json

# Setup common variables
_defaultFileName = "MovieInfo"


class WriteCSV:
    """Parse a given TheaterDetailPage.TheaterDetail object to write each show time on a single row as below:
       <theaterName>, <movieName>, <movieImageURL>, <showTime>
       Expect multiple rows per theater / movie combination
    """
    defaultFileExtension = ".csv"
    _csvHeader = ['TheaterName', 'MovieName', 'MovieImageURL', 'ShowTime']

    @staticmethod
    def write_movie_details(theaterDetail, fileName=_defaultFileName+defaultFileExtension):
        with open(fileName, 'a', newline='') as outFile:
            csvFile = csv.writer(outFile, delimiter=',', quoting=csv.QUOTE_MINIMAL)
            # Check for empty file and write headers on first line if empty
            if os.path.getsize(fileName) == 0:
                csvFile.writerow(WriteCSV._csvHeader)
            for movieDetail in theaterDetail.get_movies_list():
                for showTime in movieDetail.get_movie_show_times():
                    csvFile.writerow([theaterDetail.theaterName, movieDetail.movieName, movieDetail.movieImageURL,
                                      showTime])


# class WriteExcel:
#     """Parse a given TheaterDetailPage.TheaterDetail object to write each show time on a single row as below:
#        <theaterName>, <movieName>, <movieImageURL>, <showTime>
#        Expect multiple rows per theater / movie combination
#     """
#     defaultFileExtension = ".xls"
#     _excelHeader = ['TheaterName', 'MovieName', 'MovieImageURL', 'ShowTime']
#
#     @staticmethod
#     def write_movie_details(theaterDetail, fileName=_defaultFileName+defaultFileExtension, row=0):
#         # Setup Excel data objects
#         wb = xlwt.Workbook()
#         ws = wb.add_sheet('Movie Info')
#
#         # Check for empty file and write headers on first line if empty
#         try:
#             # Use try to determine if file does not exist, use if inside try to determine if file exists with no data.
#             if os.path.getsize(fileName) == 0:
#                 pass
#         except FileNotFoundError:
#             pass
#         finally:
#             ws.write(row, 0, WriteExcel._excelHeader[0])
#             ws.write(row, 1, WriteExcel._excelHeader[1])
#             ws.write(row, 2, WriteExcel._excelHeader[2])
#             ws.write(row, 3, WriteExcel._excelHeader[3])
#             row += 1
#
#         for movieDetail in theaterDetail.get_movies_list():
#                 for showTime in movieDetail.get_movie_show_times():
#                     ws.write(row, 0, theaterDetail.theaterName)
#                     ws.write(row, 1, movieDetail.movieName)
#                     ws.write(row, 2, movieDetail.movieImageURL)
#                     ws.write(row, 3, showTime)
#                     row += 1
#
#         wb.save(fileName)
#         return row

class WriteExcel:
    """Parse a given TheaterDetailPage.TheaterDetail object to write each show time on a single row as below:
       <theaterName>, <movieName>, <movieImageURL>, <showTime>
       Expect multiple rows per theater / movie combination
    """
    defaultFileExtension = ".xlsx"
    _excelHeader = ['TheaterName', 'MovieName', 'MovieImageURL', 'ShowTime']

    @staticmethod
    def write_movie_details(theaterDetail, fileName=_defaultFileName+defaultFileExtension):
        # Check for empty file and write headers on first line if empty
        try:
            # Setup Excel data objects
            wb = Workbook()
            # Use try to determine if file does not exist, use if inside try to determine if file exists with no data.
            if os.path.getsize(fileName) == 0:
                ws = wb.active
                ws.append(WriteExcel._excelHeader)
                wb.save(fileName)
        except FileNotFoundError:
            ws = wb.active
            ws.append(WriteExcel._excelHeader)
            wb.save(fileName)

        # Load file that now exists and write data to it
        wb = load_workbook(filename = fileName)
        ws = wb.active
        for movieDetail in theaterDetail.get_movies_list():
                for showTime in movieDetail.get_movie_show_times():
                    ws.append([theaterDetail.theaterName, movieDetail.movieName, movieDetail.movieImageURL, showTime])

        wb.save(fileName)


class WriteJSON:
    """Parse a given TheaterDetailPage.TheaterDetail object to write each show time on a single row as below:
       <theaterName>, <movieName>, <movieImageURL>, <showTime>
       Expect multiple rows per theater / movie combination
    """
    defaultFileExtension = ".txt"

    @staticmethod
    def write_movie_details(theaterDetail, fileName=_defaultFileName+defaultFileExtension):
        # Create first part of JSON string with theater name information
        info = '{"theaterInfo": { "name": "'+theaterDetail.theaterName+'",\n'
        for movieDetail in theaterDetail.get_movies_list():
            # Create second part of JSON string with movie name and URL information
            info += '"movieInfo": { "name": "'+movieDetail.movieName
            info += '",\n "src": "'+movieDetail.movieImageURL+'",\n'
            # Create third part of JSON string with movie show times, then close third part
            info += '"showTimeInfo": {'+json.dumps(movieDetail.get_movie_show_times())+'}\n'
            # Close second part of JSON string
            info += '},\n'

        # Remove trailing comma and newline from second part of JSON string, then close first part of JSON string
        info = info.rstrip(',\n')+'\n}}\n\n'

        with open(fileName, 'a', newline='') as outFile:
            outFile.write(info)

    # @staticmethod
    # def write_movie_details(theaterDetail, fileName=_defaultFileName+defaultFileExtension):
    #     t_info = {"theaterInfo": ""}
    #     t_info.update({"name": theaterDetail.theaterName})
    #     jsonInfo = json.dumps(t_info, indent=4)
    #
    #     for movieDetail in theaterDetail.get_movies_list():
    #         m_info = {"movieInfo": ""}
    #         m_info.update({movieDetail.movieName: movieDetail.movieImageURL})
    #         m_info.update({"showTimeInfo": movieDetail.get_movie_show_times()})
    #         jsonInfo += json.dumps(m_info, indent=4)
    #
    #     with open(fileName, 'a', newline='') as outFile:
    #         outFile.write(jsonInfo)
